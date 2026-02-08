"""
Roadmap: reads roadmap.csv and releases.csv, produces a timeline and exports to
Excel and draw.io.
"""

import pandas as pd
from datetime import datetime, date
from pathlib import Path
from xml.etree import ElementTree as ET
import argparse

# Roadmap CSV: one row per TASK. Columns: domain, feature, task, start_date, end_date, notes (optional), flag (optional).
# flag: "baseline" | "optional" | empty → visual accent on task bar (e.g. EKS=baseline, Rancher=optional).
# Legacy: domain, subdomain, feature, ... treated as feature=subdomain, task=feature (one bar per row).
RELEASES_COLUMNS = ["release", "start", "end"]
FLAG_VALUES = ("baseline", "optional")  # empty/null = no flag

X_START = "2025-12-01"
X_END = "2027-12-31"

# Shared palette for Excel and draw.io so visuals match
DOMAIN_COLOR_PALETTE = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
]

# Lightness factors per feature within a domain (0 = base, 1+ = lighter). New features get next index.
FEATURE_SHADE_FACTORS = [0.0, 0.15, 0.28, 0.38, 0.48, 0.56, 0.64, 0.72]


def _hex_to_rgb(hex_str: str) -> tuple[int, int, int]:
    h = hex_str.lstrip("#")
    if len(h) == 6:
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return 128, 128, 128


def _rgb_to_hex(r: int, g: int, b: int) -> str:
    return f"#{max(0, min(255, r)):02x}{max(0, min(255, g)):02x}{max(0, min(255, b)):02x}"


def _lighten_hex(hex_str: str, factor: float) -> str:
    """Lighten a hex color; factor 0 = same, 0.5 = 50% toward white."""
    r, g, b = _hex_to_rgb(hex_str)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return _rgb_to_hex(r, g, b)


def _feature_color_map(roadmap: pd.DataFrame):
    """Build (domain, feature) -> hex color. Each domain gets a base; each feature gets a shade (dynamic)."""
    df = _prepare_roadmap(roadmap)
    domains = df["domain"].unique().tolist()
    domain_base = {d: DOMAIN_COLOR_PALETTE[i % len(DOMAIN_COLOR_PALETTE)] for i, d in enumerate(domains)}
    out = {}
    feature_index_per_domain = {}  # domain -> { feature -> index }
    for domain_name, feature_name, _ in _roadmap_by_domain_feature(roadmap):
        key = (domain_name, feature_name)
        if key not in out:
            if domain_name not in feature_index_per_domain:
                feature_index_per_domain[domain_name] = {}
            if feature_name not in feature_index_per_domain[domain_name]:
                feature_index_per_domain[domain_name][feature_name] = len(feature_index_per_domain[domain_name])
            idx = feature_index_per_domain[domain_name][feature_name]
            factor = FEATURE_SHADE_FACTORS[idx % len(FEATURE_SHADE_FACTORS)]
            out[key] = _lighten_hex(domain_base[domain_name], factor)
    return out


def _month_range(x_start: str, x_end: str):
    """Yield (datetime, month_key) for each month from x_start to x_end."""
    start = datetime.strptime(x_start, "%Y-%m-%d")
    end = datetime.strptime(x_end, "%Y-%m-%d")
    d = start.replace(day=1)
    while d <= end:
        yield d, d.strftime("%b %Y")
        new_m = d.month + 1
        new_y = d.year + (new_m - 1) // 12
        new_m = (new_m - 1) % 12 + 1
        try:
            d = d.replace(year=new_y, month=new_m)
        except ValueError:
            d = d.replace(year=new_y, month=min(new_m, 12), day=1)


def normalize_columns(df: pd.DataFrame, expected: list) -> pd.DataFrame:
    """Normalize column names to lowercase and strip; map common variants."""
    df = df.copy()
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    return df


def parse_date(s) -> datetime | None:
    if pd.isna(s):
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m", "%b %Y", "%B %Y"):
        try:
            return datetime.strptime(s[:10] if len(s) >= 10 else s, fmt[: len(s)])
        except (ValueError, TypeError):
            continue
    try:
        return pd.to_datetime(s)
    except Exception:
        return None


def _ensure_date_columns(df: pd.DataFrame, start_aliases=("start", "start_date"), end_aliases=("end", "end_date")) -> pd.DataFrame:
    """Map start_date/end_date to start/end if present."""
    cols = [str(c).strip() for c in df.columns]
    for a in start_aliases:
        if a in df.columns:
            df["start"] = df[a].copy()
            break
    else:
        for i, c in enumerate(cols):
            if "start" in c.lower() and "date" in c.lower():
                df["start"] = df.iloc[:, i].copy()
                break
    for a in end_aliases:
        if a in df.columns:
            df["end"] = df[a].copy()
            break
    else:
        for i, c in enumerate(cols):
            if "end" in c.lower() and "date" in c.lower():
                df["end"] = df.iloc[:, i].copy()
                break
    return df


def _expand_single_column_csv(df: pd.DataFrame) -> pd.DataFrame:
    """If CSV was read as one column (whole line quoted), split into columns."""
    if len(df.columns) != 1:
        return df
    col = df.columns[0]
    if "domain" not in col.lower() and "start" not in col.lower():
        return df
    rows = df[col].astype(str).str.split(",", expand=True)
    # Header is in the column name (first line of file), not in rows.iloc[0] (first data line)
    header_str = str(col).strip(' "')
    if "domain" in header_str.lower() or "start_date" in header_str:
        names = [s.strip(' "') for s in header_str.split(",")]
        # Pad or trim to match column count
        n = rows.shape[1]
        if len(names) < n:
            names.extend(f"_col{i}" for i in range(len(names), n))
        else:
            names = names[:n]
        rows.columns = names
        # No row to drop; first line was header and is already the column name
    return rows


def load_roadmap(path: str = "roadmap.csv") -> pd.DataFrame:
    df = pd.read_csv(path)
    df = _expand_single_column_csv(df)
    df = normalize_columns(df, [])
    df = _ensure_date_columns(df)
    if "start" not in df.columns or "end" not in df.columns:
        raise ValueError(
            f"Roadmap CSV must have start/end or start_date/end_date columns. Got: {list(df.columns)}"
        )
    for col in ["start", "end"]:
        df[col] = df[col].apply(parse_date)
    df = df.dropna(subset=["start", "end"])
    return df


def load_releases(path: str = "releases.csv") -> pd.DataFrame:
    df = pd.read_csv(path)
    df = _expand_single_column_csv(df)
    df = normalize_columns(df, RELEASES_COLUMNS)
    df = _ensure_date_columns(df)
    if "start" not in df.columns or "end" not in df.columns:
        raise ValueError(
            f"Releases CSV must have start/end or start_date/end_date columns. Got: {list(df.columns)}"
        )
    for col in ["start", "end"]:
        df[col] = df[col].apply(parse_date)
    df = df.dropna(subset=["start", "end"])
    return df


def _prepare_roadmap(roadmap: pd.DataFrame) -> pd.DataFrame:
    """Normalize and add row_feature (chart row) and task_label (bar label). Supports:
    - New format: domain, feature, task -> row_feature=feature, task_label=task (multiple tasks per feature).
    - Legacy: domain, subdomain, feature -> row_feature=subdomain or domain, task_label=feature (one bar per row).
    """
    df = roadmap.copy()
    df["domain"] = df.get("domain", "").fillna("").astype(str).str.strip()
    for col in ("subdomain", "feature", "task"):
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()
    # Canonical row (feature) and task (bar) labels
    has_task = "task" in df.columns and df["task"].astype(str).str.strip().str.len().gt(0).any()
    if has_task:
        df["row_feature"] = df["feature"]
        df["task_label"] = df["task"].astype(str).str.strip().replace("", pd.NA).fillna(df["feature"])
    else:
        # Legacy: subdomain = row, feature = single task
        df["row_feature"] = df.get("subdomain", "").replace("", pd.NA).fillna(df["domain"])
        df["task_label"] = df["feature"]
    df["row_feature"] = df["row_feature"].astype(str)
    df["task_label"] = df["task_label"].astype(str)
    # Normalize optional flag: baseline | optional | empty
    if "flag" in df.columns:
        f = df["flag"].astype(str).str.strip().str.lower()
        df["flag"] = f.where(f.isin(FLAG_VALUES), None).replace("", None).replace("nan", None)
    else:
        df["flag"] = None
    # Order by first appearance in CSV: domain order, then feature order, then start date
    domain_order = {d: i for i, d in enumerate(df["domain"].drop_duplicates())}
    df["_domain_ord"] = df["domain"].map(domain_order)
    feature_key = df["domain"] + "\0" + df["row_feature"]
    feature_order = {k: i for i, k in enumerate(feature_key.drop_duplicates())}
    df["_feature_ord"] = feature_key.map(feature_order)
    df = df.sort_values(["_domain_ord", "_feature_ord", "start"]).drop(columns=["_domain_ord", "_feature_ord"])
    return df


def _roadmap_by_domain_feature(roadmap: pd.DataFrame):
    """Yield (domain, feature, list of task dicts) for chart: one row per feature, multiple tasks = multiple bars.
    Each task dict has: start, end, task_label, release, status, priority, notes, etc.
    """
    df = _prepare_roadmap(roadmap)
    domain, feature, group = None, None, []
    for _, row in df.iterrows():
        r = row.to_dict()
        d, f = r.get("domain"), r.get("row_feature")
        if (d, f) != (domain, feature):
            if group:
                yield domain, feature, group
            domain, feature, group = d, f, [r]
        else:
            group.append(r)
    if group:
        yield domain, feature, group


def _roadmap_by_domain(roadmap: pd.DataFrame):
    """Yield (domain_name, list of task row dicts) for Excel flat list. Each row = one task."""
    df = _prepare_roadmap(roadmap)
    domain = None
    group = []
    for _, row in df.iterrows():
        r = row.to_dict()
        if r.get("domain") != domain:
            if group:
                yield domain, group
            domain = r.get("domain")
            group = [r]
        else:
            group.append(r)
    if group:
        yield domain, group


def _merge_consecutive_cells(ws, row_keys, col, start_row, alignment=None):
    """Merge consecutive rows in column col where row_keys[i] equals row_keys[i-1].
    row_keys: list of values (one per row from start_row). Merge when value same as previous."""
    if alignment is None:
        try:
            from openpyxl.styles import Alignment
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        except ImportError:
            return
    r = start_row
    i = 0
    while i < len(row_keys):
        j = i + 1
        while j < len(row_keys) and row_keys[j] == row_keys[i]:
            j += 1
        if j > i + 1:
            ws.merge_cells(start_row=r, start_column=col, end_row=r + (j - i) - 1, end_column=col)
            ws.cell(r, col).alignment = alignment
        r += j - i
        i = j


def _month_year_str(d) -> str:
    """Format date as Month Year only (no time)."""
    if d is None or (hasattr(d, "year") and pd.isna(d)):
        return ""
    if hasattr(d, "strftime"):
        return d.strftime("%b %Y")
    return str(d)[:7]  # fallback

def export_to_excel(roadmap: pd.DataFrame, releases: pd.DataFrame, path: Path) -> None:
    """Export roadmap and releases to Excel: data sheets (Month/Year dates) + visual Timeline sheet."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        raise ImportError("Excel export requires openpyxl: pip install openpyxl") from None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- Sheet 1: Visual Roadmap (timeline grid) - primary visual ----
    months_list = list(_month_range(X_START, X_END))
    feature_to_color = _feature_color_map(roadmap)
    # Only horizontal borders: bold under domain, thin under feature. Task bars have no border (no random outlines).
    no_border = Border()
    bold_bottom = Side(style="medium")
    thin_bottom = Side(style="thin")
    flag_baseline_left = Side(style="medium", color="2E7D32")   # green accent for baseline (e.g. EKS)
    flag_optional_left = Side(style="mediumDashed", color="EF6C00")   # orange bold dashed for optional (e.g. Rancher)

    ws_vis = wb.create_sheet("Timeline", 0)
    domain_col, feature_col, task_col = 1, 2, 3
    label_col_w = 4  # months start at column 4
    n_months = len(months_list)
    # Header: Domain | Feature | Task | Dec 2025 | ...
    ws_vis.cell(1, 1, "Roadmap (Dec 2025 – Dec 2027)")
    ws_vis.cell(1, 1).font = Font(bold=True, size=12)
    ws_vis.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws_vis.cell(2, 1, "Domain")
    ws_vis.cell(2, 2, "Feature")
    ws_vis.cell(2, 3, "Task")
    for c in (1, 2, 3):
        ws_vis.cell(2, c).font = Font(bold=True, size=9)
    for c, (_, month_label) in enumerate(months_list, start=label_col_w):
        ws_vis.cell(2, c, month_label)
        ws_vis.cell(2, c).font = Font(bold=True, size=9)
        ws_vis.cell(2, c).alignment = Alignment(horizontal="center", wrap_text=True)
    # Release row (span label cols)
    ws_vis.cell(3, 1, "Releases")
    ws_vis.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    ws_vis.cell(3, 1).font = Font(bold=True, size=9)
    for c, (month_dt, _) in enumerate(months_list, start=label_col_w):
        month_start = pd.Timestamp(month_dt)
        month_end = month_start + pd.offsets.MonthEnd(0)
        for _, rel in releases.iterrows():
            s, e = rel.get("start"), rel.get("end")
            if s is None or e is None:
                continue
            start_ts, end_ts = pd.Timestamp(s), pd.Timestamp(e)
            if month_start <= end_ts and start_ts <= month_end:
                ws_vis.cell(3, c, rel.get("release", ""))
                ws_vis.cell(3, c).fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
                ws_vis.cell(3, c).font = Font(size=8)
                break
    # Build flat list: one row per task (domain, feature, task_label, start, end, ...)
    flat_rows = []
    for domain_name, feature_name, task_list in _roadmap_by_domain_feature(roadmap):
        for r in task_list:
            flat_rows.append((domain_name, feature_name, r))
    # Detect last row of each feature and each domain for borders
    is_last_of_feature = []
    is_last_of_domain = []
    prev_d, prev_f = None, None
    for i in range(len(flat_rows) - 1, -1, -1):
        d, f, _ = flat_rows[i]
        is_last_of_domain.append(i == len(flat_rows) - 1 or flat_rows[i + 1][0] != d)
        is_last_of_feature.append(i == len(flat_rows) - 1 or flat_rows[i + 1][0] != d or flat_rows[i + 1][1] != f)
    is_last_of_domain.reverse()
    is_last_of_feature.reverse()
    row_idx = 4
    for idx, (domain_name, feature_name, r) in enumerate(flat_rows):
        ws_vis.cell(row_idx, domain_col, domain_name)
        ws_vis.cell(row_idx, domain_col).font = Font(bold=True, size=9)
        ws_vis.cell(row_idx, feature_col, feature_name)
        ws_vis.cell(row_idx, feature_col).font = Font(size=9)
        ws_vis.cell(row_idx, task_col, r.get("task_label", ""))
        ws_vis.cell(row_idx, task_col).font = Font(size=9)
        # Label columns: no border on task rows; only bottom line on separator rows
        if is_last_of_domain[idx]:
            for col in range(1, label_col_w):
                ws_vis.cell(row_idx, col).border = Border(bottom=bold_bottom)
        elif is_last_of_feature[idx]:
            for col in range(1, label_col_w):
                ws_vis.cell(row_idx, col).border = Border(bottom=thin_bottom)
        else:
            for col in range(1, label_col_w):
                ws_vis.cell(row_idx, col).border = no_border
        # Task bar cells: no border inside the bar; only the leftmost cell gets flag accent; separator rows get bottom only
        flag = r.get("flag") or ""
        left_side = flag_baseline_left if flag == "baseline" else (flag_optional_left if flag == "optional" else None)
        bottom_side = bold_bottom if is_last_of_domain[idx] else (thin_bottom if is_last_of_feature[idx] else None)
        color = feature_to_color.get((domain_name, feature_name), "CCCCCC")
        fill_hex = color.lstrip("#")[:6]
        first_filled_col = None
        for c, (month_dt, _) in enumerate(months_list, start=label_col_w):
            start_ts = pd.Timestamp(r["start"])
            end_ts = pd.Timestamp(r["end"])
            month_start = pd.Timestamp(month_dt)
            month_end = month_start + pd.offsets.MonthEnd(0)
            if month_start <= end_ts and start_ts <= month_end:
                cell = ws_vis.cell(row_idx, c, "")
                cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
                if first_filled_col is None:
                    first_filled_col = c
                # Only the leftmost cell of the bar gets the flag accent; no other borders on the bar
                cell.border = Border(left=left_side if first_filled_col == c else None, bottom=bottom_side)
        row_idx += 1
    # Merge Domain and Feature for consecutive same values (like the reference screenshot)
    _merge_consecutive_cells(ws_vis, [d for d, _, _ in flat_rows], domain_col, 4)
    _merge_consecutive_cells(ws_vis, [(d, f) for d, f, _ in flat_rows], feature_col, 4)

    # Re-apply borders after merge: flag accent only on leftmost cell of each task bar; no borders inside bars
    r = 4
    for idx in range(len(flat_rows)):
        _, _, row_dict = flat_rows[idx]
        flag = row_dict.get("flag") or ""
        left_side = flag_baseline_left if flag == "baseline" else (flag_optional_left if flag == "optional" else None)
        bottom_side = bold_bottom if is_last_of_domain[idx] else (thin_bottom if is_last_of_feature[idx] else None)
        start_ts = pd.Timestamp(row_dict["start"])
        end_ts = pd.Timestamp(row_dict["end"])
        for col in range(1, label_col_w):
            ws_vis.cell(r, col).border = Border(bottom=bottom_side) if bottom_side else no_border
        first_filled = None
        for c, (month_dt, _) in enumerate(months_list, start=label_col_w):
            month_start = pd.Timestamp(month_dt)
            month_end = month_start + pd.offsets.MonthEnd(0)
            if month_start <= end_ts and start_ts <= month_end:
                if first_filled is None:
                    first_filled = c
                ws_vis.cell(r, c).border = Border(left=left_side if c == first_filled else None, bottom=bottom_side)
            else:
                ws_vis.cell(r, c).border = Border(bottom=bottom_side) if bottom_side else no_border
        r += 1

    # Red "today" column: highlight current month with a vertical line (red right border)
    today = date.today()
    current_month_col = None
    for c, (month_dt, _) in enumerate(months_list, start=label_col_w):
        if month_dt.month == today.month and month_dt.year == today.year:
            current_month_col = c
            break
    if current_month_col is not None:
        red_side = Side(style="medium", color="D32F2F")
        for r in range(2, row_idx):
            cell = ws_vis.cell(r, current_month_col)
            b = cell.border
            cell.border = Border(
                left=b.left, right=red_side, top=b.top, bottom=b.bottom
            )

    # Hide gridlines so only our explicit borders show (no random lines inside task bars)
    try:
        ws_vis.sheet_view.showGridLines = False
    except Exception:
        pass

    # Legend: Baseline vs Optional (below the timeline data)
    legend_row = row_idx + 2
    ws_vis.cell(legend_row, 1, "Legend:")
    ws_vis.cell(legend_row, 1).font = Font(bold=True, size=9)
    ws_vis.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=3)
    legend_row += 1
    ws_vis.cell(legend_row, 1, "Baseline (primary offering)")
    ws_vis.cell(legend_row, 1).font = Font(size=9)
    ws_vis.cell(legend_row, 1).border = Border(left=flag_baseline_left)
    ws_vis.cell(legend_row, 2, "Green left accent")
    ws_vis.cell(legend_row, 2).font = Font(size=9)
    ws_vis.cell(legend_row, 3, "Optional")
    ws_vis.cell(legend_row, 3).font = Font(size=9)
    ws_vis.cell(legend_row, 3).border = Border(left=flag_optional_left)
    ws_vis.cell(legend_row, 4, "Orange dashed left accent")
    ws_vis.cell(legend_row, 4).font = Font(size=9)

    ws_vis.column_dimensions["A"].width = 20
    ws_vis.column_dimensions["B"].width = 18
    ws_vis.column_dimensions["C"].width = 32
    for c in range(4, n_months + label_col_w):
        ws_vis.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 8
    ws_vis.freeze_panes = "D4"

    # ---- Sheet 2: Roadmap data (Month/Year for dates) ----
    ws_rm = wb.create_sheet("Roadmap", 1)
    df_rm = _prepare_roadmap(roadmap)
    out_cols = ["domain", "row_feature", "task_label", "start", "end", "notes", "flag"]
    cols = [c for c in out_cols if c in df_rm.columns]
    df_rm = df_rm[cols].copy()
    df_rm = df_rm.rename(columns={"row_feature": "feature", "task_label": "task"})
    df_rm["start"] = df_rm["start"].apply(_month_year_str)
    df_rm["end"] = df_rm["end"].apply(_month_year_str)
    for r_idx, row in enumerate(dataframe_to_rows(df_rm, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_rm.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
    for cell in ws_rm[1]:
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col in ws_rm.columns:
        ws_rm.column_dimensions[col[0].column_letter].width = 14

    # ---- Sheet 3: Releases data (Month/Year for dates) ----
    ws_rel = wb.create_sheet("Releases", 2)
    ws_rel.cell(1, 1, "release")
    ws_rel.cell(1, 2, "start")
    ws_rel.cell(1, 3, "end")
    for cell in ws_rel[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for r_idx, row in releases.iterrows():
        r_num = r_idx + 2
        ws_rel.cell(r_num, 1, row.get("release", ""))
        ws_rel.cell(r_num, 2, _month_year_str(row.get("start")))
        ws_rel.cell(r_num, 3, _month_year_str(row.get("end")))
    ws_rel.column_dimensions["A"].width = 10
    ws_rel.column_dimensions["B"].width = 12
    ws_rel.column_dimensions["C"].width = 12

    wb.save(path)
    print(f"Exported Excel: {path}")


def export_to_drawio(roadmap: pd.DataFrame, releases: pd.DataFrame, path: Path) -> None:
    """Export a timeline diagram to draw.io XML (.drawio)."""
    x_min = datetime.strptime(X_START, "%Y-%m-%d")
    x_max = datetime.strptime(X_END, "%Y-%m-%d")
    total_days = (x_max - x_min).days or 1
    timeline_width = 900
    row_height = 32
    margin_left = 40
    margin_top = 60
    band_height = 24  # release band lane height
    domain_col_w = 180   # domain group box width (wider so names wrap and fit)
    feature_col_w = 220  # feature label column (wider so names wrap and fit)
    label_width = domain_col_w + feature_col_w + 20  # total left side

    df = _prepare_roadmap(roadmap)
    feature_to_color = _feature_color_map(roadmap)

    def date_to_x(d):
        if d is None:
            return 0
        days = (d - x_min).days
        return margin_left + label_width + (days / total_days) * timeline_width

    def date_to_w(start, end):
        if start is None or end is None:
            return 20
        return max(4, (end - start).days / total_days * timeline_width)

    # Build XML: draw.io uses mxfile containing mxGraphModel
    root = ET.Element("mxfile", host="app.diagrams.net")
    root.set("modified", datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.000Z"))
    root.set("agent", "Python")
    root.set("version", "22.1.0")
    root.set("etag", "roadmap")
    root.set("type", "device")

    diagram = ET.SubElement(root, "diagram", id="roadmap", name="Roadmap")
    mx = ET.SubElement(diagram, "mxGraphModel", dx="1200", dy="800", grid="1", gridSize="10",
                       guides="1", tooltips="1", connect="1", arrows="1", fold="1", page="1",
                       pageScale="1", pageWidth="1400", pageHeight="1200", math="0", shadow="0")
    mx_root = ET.SubElement(mx, "root")
    ET.SubElement(mx_root, "mxCell", id="0")
    ET.SubElement(mx_root, "mxCell", id="1", parent="0")

    def add_cell(parent_id, cell_id, value="", style="", x=0, y=0, w=80, h=24, vertex="1"):
        cell = ET.SubElement(mx_root, "mxCell", id=str(cell_id), parent=parent_id, vertex=vertex)
        if value:
            cell.set("value", str(value)[:200])
        if style:
            cell.set("style", style)
        geom = ET.SubElement(cell, "mxGeometry", x=str(round(x, 1)), y=str(round(y, 1)),
                             width=str(round(w, 1)), height=str(round(h, 1)))
        geom.set("as", "geometry")
        return cell

    cell_id = 10

    # Content bounds – one row per task (same as Excel) so draw.io stays synced with domains
    flat_list = []
    for domain_name, feature_name, task_list in _roadmap_by_domain_feature(roadmap):
        for r in task_list:
            flat_list.append((domain_name, feature_name, r))
    n_data_rows = len(flat_list)
    y_header = margin_top
    content_bottom = margin_top + 28 + n_data_rows * row_height
    content_right = margin_left + label_width + timeline_width
    pad = 24
    legend_h = 56  # space for legend below timeline

    # 1) Background panel – rectangle (no rounded corners) containing the whole roadmap + legend
    bg_x = margin_left - 24
    bg_y = 0
    bg_w = content_right - margin_left + pad + 24
    bg_h = content_bottom + pad + legend_h
    add_cell("1", cell_id, "",
             "rounded=0;whiteSpace=wrap;fillColor=#fafafa;strokeColor=#bdbdbd;strokeWidth=2;",
             bg_x, bg_y, bg_w, bg_h)
    cell_id += 1

    # 2) Light vertical guide lines at each release start and end
    for _, row in releases.iterrows():
        start, end = row.get("start"), row.get("end")
        if start is None:
            continue
        for d in (start, end):
            if d is None:
                continue
            x = date_to_x(d)
            if x < margin_left + label_width or x > content_right:
                continue
            add_cell("1", cell_id, "",
                     "fillColor=#e0e0e0;strokeColor=none;",
                     round(x - 0.5, 1), y_header, 1, content_bottom - y_header)
            cell_id += 1

    # 3) Red "today" line – current date (clamped to roadmap range)
    today = date.today()
    if x_min.date() <= today <= x_max.date():
        x_today = date_to_x(datetime.combine(today, datetime.min.time()))
        add_cell("1", cell_id, "",
                 "rounded=0;whiteSpace=wrap;fillColor=#d32f2f;strokeColor=none;",
                 x_today - 1, y_header, 3, content_bottom - y_header)
        cell_id += 1

    # Text/label style: dark text so it's readable on any background
    font_dark = "fontColor=#1a1a1a;"
    font_dark_small = "fontColor=#333333;fontSize=9;"

    # Title (moved up so it sits clear above the bands)
    add_cell("1", cell_id, "Roadmap (Dec 2025 – Dec 2027)",
             f"text;html=1;strokeColor=none;fillColor=#ffffff;align=left;verticalAlign=middle;fontSize=14;fontStyle=1;{font_dark}",
             margin_left, 4, 500, 24)
    cell_id += 1

    # Release bands (background lane) – light gray with dark text (span both label columns)
    y_band = margin_top - band_height - 8
    add_cell("1", cell_id, "Releases",
             f"rounded=1;whiteSpace=wrap;fillColor=#e8e8e8;strokeColor=#666666;fontStyle=1;{font_dark}",
             margin_left, y_band, domain_col_w + feature_col_w, band_height)
    cell_id += 1
    for _, row in releases.iterrows():
        start, end = row.get("start"), row.get("end")
        if start is None or end is None:
            continue
        x, w = date_to_x(start), date_to_w(start, end)
        add_cell("1", cell_id, str(row.get("release", "")),
                 f"rounded=1;whiteSpace=wrap;fillColor=#d0d0d0;strokeColor=#888888;{font_dark}",
                 x, y_band, w, band_height)
        cell_id += 1

    # Timeline header (month ticks) – light background, dark text
    y_header = margin_top
    add_cell("1", cell_id, "", "whiteSpace=wrap;fillColor=#e5e5e5;strokeColor=#999999;",
             margin_left + label_width, y_header, timeline_width, 22)
    cell_id += 1
    # Month labels (every month) – narrow width so all fit
    d = x_min
    month_w = max(18, timeline_width / 25)  # ~25 months
    while d <= x_max:
        x = date_to_x(d) - month_w / 2
        add_cell("1", cell_id, d.strftime("%b %y"),
                 f"text;html=1;strokeColor=none;fillColor=none;align=center;{font_dark_small}",
                 x, y_header + 2, month_w, 18)
        cell_id += 1
        new_m = d.month + 1
        new_y = d.year + (new_m - 1) // 12
        new_m = (new_m - 1) % 12 + 1
        try:
            d = d.replace(year=new_y, month=new_m)
        except ValueError:
            d = d.replace(year=new_y, month=min(new_m, 12), day=1)

    # Domain blocks: (domain_name, start_y, n_rows) for drawing merged domain boxes
    y = margin_top + 28
    domain_blocks = []
    i = 0
    while i < len(flat_list):
        domain_name = flat_list[i][0]
        start_y = y
        n_rows = 0
        while i < len(flat_list) and flat_list[i][0] == domain_name:
            n_rows += 1
            y += row_height
            i += 1
        domain_blocks.append((domain_name, start_y, n_rows))

    # Feature blocks: (feature_name, start_row_index, n_rows) so we draw one label per feature (no repeat)
    feature_blocks = []
    i = 0
    while i < len(flat_list):
        domain_name, feature_name = flat_list[i][0], flat_list[i][1]
        start_i = i
        n_rows = 0
        while i < len(flat_list) and flat_list[i][0] == domain_name and flat_list[i][1] == feature_name:
            n_rows += 1
            i += 1
        feature_blocks.append((domain_name, feature_name, start_i, n_rows))

    # Draw domain group boxes first (behind) – each spans all its task rows
    for domain_name, y_start, n_rows in domain_blocks:
        box_h = n_rows * row_height
        add_cell("1", cell_id, "",
                 f"rounded=1;whiteSpace=wrap;fillColor=#f0f0f0;strokeColor=#999999;{font_dark}",
                 margin_left, y_start, domain_col_w, box_h)
        cell_id += 1
        add_cell("1", cell_id, (domain_name or "").upper(),
                 f"text;html=1;whiteSpace=wrap;strokeColor=none;fillColor=none;align=center;verticalAlign=middle;fontSize=11;fontStyle=1;{font_dark}",
                 margin_left + 4, y_start + 4, domain_col_w - 8, box_h - 8)
        cell_id += 1

    # Feature column: one label per feature block (grouped, like Excel) – no repeating feature name per task
    for _domain, feature_name, start_i, n_rows in feature_blocks:
        y_start = margin_top + 28 + start_i * row_height
        box_h = n_rows * row_height
        add_cell("1", cell_id, (feature_name or "").strip(),
                 f"text;html=1;whiteSpace=wrap;strokeColor=none;fillColor=#ffffff;align=left;verticalAlign=middle;fontSize=10;{font_dark}",
                 margin_left + domain_col_w + 4, y_start + 4, feature_col_w - 8, box_h - 8)
        cell_id += 1

    # One row per task: one bar per row; wrap text so labels fit, dynamic font size
    bar_h = 22
    y = margin_top + 28
    for domain_name, feature_name, r in flat_list:
        color = feature_to_color.get((domain_name, feature_name), "#cccccc")
        x, w = date_to_x(r["start"]), date_to_w(r["start"], r["end"])
        bar_w = max(w, 24)
        raw_label = (r.get("task_label") or "").strip()
        # Allow wrap so full label can use 2 lines; only truncate if very long (keep under 50 chars for readability)
        label = raw_label if len(raw_label) <= 50 else raw_label[:47] + "…"
        # Smaller font for narrow bars so more text fits when wrapped
        if bar_w < 45:
            font_size = 6
        elif bar_w < 75:
            font_size = 7
        else:
            font_size = 8
        flag = r.get("flag") or ""
        stroke_style = "strokeColor=#333333;"
        if flag == "optional":
            stroke_style = "strokeColor=#EF6C00;dashPattern=1 2 1 2;"
        style = f"rounded=1;whiteSpace=wrap;fillColor={color};{stroke_style}fontColor=#ffffff;fontSize={font_size};align=left;verticalAlign=middle;spacingLeft=3;spacingRight=2;overflow=hidden;"
        add_cell("1", cell_id, label, style, x, y + (row_height - bar_h) / 2, bar_w, bar_h)
        cell_id += 1
        # Baseline flag: green vertical accent at left of bar
        if flag == "baseline":
            accent_w = min(5, bar_w // 4)
            add_cell("1", cell_id, "", "rounded=0;whiteSpace=wrap;fillColor=#2E7D32;strokeColor=none;",
                     x, y + (row_height - bar_h) / 2, accent_w, bar_h)
            cell_id += 1
        y += row_height

    # Legend (below timeline): Baseline vs Optional
    y_legend = content_bottom + 12
    add_cell("1", cell_id, "Legend:",
             f"text;html=1;strokeColor=none;fillColor=none;align=left;verticalAlign=middle;fontSize=11;fontStyle=1;{font_dark}",
             margin_left, y_legend, 60, 18)
    cell_id += 1
    y_legend += 22
    # Baseline: green accent + label
    add_cell("1", cell_id, "", "rounded=0;whiteSpace=wrap;fillColor=#2E7D32;strokeColor=none;",
             margin_left, y_legend + 4, 8, 14)
    cell_id += 1
    add_cell("1", cell_id, "Baseline (primary offering)",
             f"text;html=1;strokeColor=none;fillColor=none;align=left;verticalAlign=middle;fontSize=10;{font_dark}",
             margin_left + 14, y_legend, 200, 22)
    cell_id += 1
    # Optional: orange dashed bar + label
    add_cell("1", cell_id, "", "rounded=1;whiteSpace=wrap;fillColor=#FFF3E0;strokeColor=#EF6C00;dashPattern=1 2 1 2;strokeWidth=1;",
             margin_left + 220, y_legend + 6, 24, 10)
    cell_id += 1
    add_cell("1", cell_id, "Optional",
             f"text;html=1;strokeColor=none;fillColor=none;align=left;verticalAlign=middle;fontSize=10;{font_dark}",
             margin_left + 250, y_legend, 120, 22)
    cell_id += 1

    tree = ET.ElementTree(root)
    ET.indent(tree, space="  ")
    with open(path, "wb") as f:
        f.write(b'<?xml version="1.0" encoding="UTF-8"?>\n')
        tree.write(f, encoding="utf-8", default_namespace="", method="xml")
    print(f"Exported draw.io: {path}")


def main():
    parser = argparse.ArgumentParser(description="Export roadmap from roadmap.csv and releases.csv")
    parser.add_argument("roadmap_csv", nargs="?", default="roadmap.csv", help="Path to roadmap CSV (default: roadmap.csv)")
    parser.add_argument("releases_csv", nargs="?", default="releases.csv", help="Path to releases CSV (default: releases.csv)")
    parser.add_argument("--excel", action="store_true", default=True, help="Export to Excel (default: True)")
    parser.add_argument("--no-excel", action="store_false", dest="excel", help="Skip Excel export")
    parser.add_argument("--drawio", action="store_true", default=True, help="Export to draw.io (default: True)")
    parser.add_argument("--no-drawio", action="store_false", dest="drawio", help="Skip draw.io export")
    args = parser.parse_args()

    base = Path(__file__).resolve().parent
    roadmap_path = base / args.roadmap_csv if not Path(args.roadmap_csv).is_absolute() else Path(args.roadmap_csv)
    releases_path = base / args.releases_csv if not Path(args.releases_csv).is_absolute() else Path(args.releases_csv)

    def resolve_csv(p: Path) -> Path:
        if p.exists():
            return p
        # Try with extra .csv (e.g. roadmap.csv -> roadmap.csv.csv)
        fallback = p.parent / (p.name + ".csv")
        if fallback.exists():
            return fallback
        return p

    roadmap_path = resolve_csv(roadmap_path)
    releases_path = resolve_csv(releases_path)

    if not roadmap_path.exists():
        raise FileNotFoundError(f"Missing {roadmap_path} (also looked for {roadmap_path}.csv)")
    if not releases_path.exists():
        raise FileNotFoundError(f"Missing {releases_path} (also looked for {releases_path}.csv)")

    roadmap = load_roadmap(roadmap_path)
    releases = load_releases(releases_path)

    if roadmap.empty:
        raise ValueError("roadmap.csv has no valid rows with start/end dates")

    if args.excel:
        export_to_excel(roadmap, releases, base / "roadmap.xlsx")

    if args.drawio:
        export_to_drawio(roadmap, releases, base / "roadmap.drawio")


if __name__ == "__main__":
    main()
